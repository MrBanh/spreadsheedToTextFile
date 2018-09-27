#! python3

# spreadsheetToTextFile.py - Opens a spreadsheet and write the cells of each column to a separate text file,
# and each row is a line in each textfile.

import os
import openpyxl
from openpyxl.utils import get_column_letter
import sys

desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop\\')

# Converts the text file to spreadsheet
def spreadsheetToTextFile(excelFile):
    os.chdir(desktop)
    
    # Validates if files exist
    try:
        # Loads the excel file and assigns the active sheet to a variable
        wb = openpyxl.load_workbook(excelFile)
        sheet = wb.active

        # Loop through each column
        for colNum in range(1, sheet.max_column + 1):
            # Creates a text file in write mode
            txtFile = open(f'spreadsheetToTextFile_{colNum}.txt', 'w')

            # Loops through every row in each column
            for rowNum in range(1, sheet.max_row + 1):
                # Extract the cells in each column and writes it to the text file. Each cell is a line in the text file
                txtFile.write(f'{sheet[get_column_letter(colNum) + str(rowNum)].value}\n')

            # Close the text file so we can create another text file for the next column
            txtFile.close()

    except FileNotFoundError as notFoundError:
        sys.exit(notFoundError)
    
    except Exception as err:
        sys.exit(err)
    

# Get text files via command line
if len(sys.argv) == 2:
    spreadsheetToTextFile(sys.argv[1])
else:
    print('Invalid number of arguments. Please enter as: spreadsheetToTextFile <excelfile>.xlsx')